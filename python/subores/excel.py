from openpyxl import load_workbook
fname = "menalookup.xlsx"

wb = load_workbook(fname)
ws = wb["vstup"]
ws_veky = wb["veky"]

arr = []
for row in ws.iter_rows(min_row=5, max_col=2, max_row=7):
    t = []
    for cell in row:
        t.append(cell.value)
    arr.append(t)

for i in range(len(arr)):
    ws_veky.cell(row=i+1, column=1, value=arr[i][1])

wb.save(fname)